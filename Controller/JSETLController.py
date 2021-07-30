#!/usr/bin/env python
# coding=utf-8
# ========================================
#
# Time: 2020/02/8 11:06
# Author: Sun
# Software: PyCharm
# Description:
# 负责整理 excel 抽取, 配置获取等操作的调配
#
# ========================================
import os
import threading
import tkinter
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from tkinter import messagebox

from Delegate.ConfigureDelegate import ConfigureDelegate
from Model.JSCellModel import JSCellModel
from Tools.JSExcelHandler import JSExcelHandler



class JSETLController(ConfigureDelegate):
    # def __init__(self):
    #     self.etlwindow = JSETLWindow(tkinter.Tk(), self.__init__())
    #
    # def initETLWindow(self):
    #     self.etlwindow = JSETLWindow(tkinter.Tk(), None)
    #     self.etlwindow.delegate = self

    # 自动根据数据提取表头
    def autoExtractSheetHead(self, configuremodel):
        """
        自动根据数据提取表头
        :param configuremodel: 读取 view 层的 model
        """
        print("开始抽取降范式的模板表头")
        datapath = configuremodel.datapath
        # 所有需要合并的表格数据
        exceldatalist = JSExcelHandler.getPathFromRootFolder(datapath)
        dflist = []
        for path in exceldatalist:
            try:
                readOpenXlsx, sheetnames, tempPath = JSExcelHandler().OpenXls(path)
                for sheetname in sheetnames:
                    rSheet = readOpenXlsx.sheet_by_name(sheetname)
                    # mergecells = rSheet.merged_cells
                    resultCells = rSheet.merged_cells
                    mergecells = []
                    # 添加扫描的行范围,默认为前 8 行进行表头的搜寻
                    rangerow = 8
                    # 根据原表中的 cellrange 确定合并单元格的范围, 筛选出 header 部分有合并单元格的情况
                    for index, cell in enumerate(resultCells):
                        crlow, crhign, cclow, cchigh = cell
                        if crlow < rangerow:
                            mergecells.append(cell)
                    # 默认第一行为表头 没有合并单元格的情况
                    if len(mergecells) == 0:
                        headrow = 1
                    else:
                        levelmap = self.createLevelMap(mergecells, rSheet)
                        # print(path)
                        dummy = self.createLevelTree(levelmap, rSheet)
                        # 根据层级来生成最终的 valuelist, keywords
                        # self.travel(dummy, valuelist, keywords, maxlevel)
                        # print(keywords)
                        maxlevel = max(levelmap.keys())
                        # 这里应该比较所有叶子节点的下边界
                        lowestnodes = levelmap[maxlevel]
                        # print(lowestnodes)
                        headrow = 0
                        for node in lowestnodes:
                            if len(node.child) > 0:
                                headrow = maxlevel + 2
                                # break
                            rlow, rhign, clow, chigh = node.cellrange
                            if (rhign > headrow):
                                headrow = rhign


                    # 找到每个 sheet 的表头范围进行和之前的对比 TODO 这里可以对提取出来的表头进行处理
                    curdf = pd.read_excel(path, sheetname, nrows=headrow, header=None)
                    # 每个进行遍历对比,完全不同则添加到 list 中
                    flag = True
                    for olddf in dflist:
                        if curdf.equals(olddf):
                            # JSExcelHandler.pathlog(path)
                            flag = False
                            pass
                    # 第一次
                    if len(dflist) == 0:
                        flag = True
                    if flag:
                        dflist.append(curdf)
                        filename, suffix = JSExcelHandler.SplitPathReturnNameAndSuffix(path)
                        # 这里要保存为 xlsx 为了兼容合并单元格的功能
                        objectpath = "{}/{}_{}.{}".format(configuremodel.headspath, filename, sheetname, 'xlsx')
                        self.restoreHead(path, sheetname, objectpath, curdf, headrow)
            except Exception as e:
                print(str(e))
                JSExcelHandler.errorlog("自动根据数据提取表头,原数据文件有问题-{}".format(path))
                pass
        print("抽取模板表头完成")


    # 根据 merge cell 定位多范式表格的位置, 以最小range 作为标准, 如果没有 merge cell 则默认第一行为表头
    # 应该将叶子节点的数据向根节点拼接,最后降范式的值从根节点层按列 + 1的顺序得出
    def lowerDimensionOfTitle(self, path, startrow=0):
        """
        进行降范式操作
        :param path: 目标文件的路径
        :param startrow: 开始降范式的起始行(因为会有标题的情况,不属于表头范围,需要去掉)
        :return: 最底层子树的标题, 追加数据的起始行, keywords

        """
        readOpenXlsx, sheetnames, tempPath = JSExcelHandler().OpenXls(path)
        rSheet = readOpenXlsx.sheet_by_name(sheetnames[0])
        valuelist = []
        # 表头层级最下面的的标题
        keywords= []
        lastrow = 0
        # 表格的列边界
        ecol = rSheet.ncols
        resultCells = rSheet.merged_cells
        # mergeCellCount = len(mergedCells)
        mergedCells = []
        # 这里要加从第几行开始降范式,默认是从 0 行开始, 默认是从 0 行开始
        if startrow > 0:
            for index, cell in enumerate(resultCells):
                crlow, crhign, cclow, cchigh = cell
                if crlow >= startrow:
                    mergedCells.append(cell)
        else:
            mergedCells = resultCells
        mergeCellCount = len(mergedCells)
        # 没有 mergeCell 的情况, 默认第一行为表头
        if mergeCellCount == 0:
            for colindex in range(0, ecol):
                value = rSheet.cell_value(startrow, colindex)
                valuelist.append(value)
                keywords.append(value)
            # print("判断--" + rSheet.name + "--的合并单元格已结束")
            # 参数返回：范式标识（False为一范式，True为多范式）、表头值、数据行坐标
            print(valuelist)
            lastrow = 1 + startrow
            return valuelist, lastrow, keywords
        else:
            levelmap = self.createLevelMap(mergedCells, rSheet)
            maxlevel = max(levelmap.keys())
            dummy = self.createLevelTree(levelmap, rSheet)
            # 根据层级来生成最终的 valuelist, keywords
            self.travel(dummy, valuelist, keywords, maxlevel)
            # print(keywords)
            # 这里应该比较所有叶子节点的下边界
            lowestnodes = levelmap[maxlevel]
            # print(lowestnodes)
            headrow = 0
            for node in lowestnodes:
                if len(node.child) > 0:
                    headrow = maxlevel + 2
                    # break
                rlow, rhign, clow, chigh = node.cellrange
                if (rhign > headrow):
                    headrow = rhign
            lastrow = headrow
            # lastrow = max(levelmap.keys()) + 1
            return valuelist, lastrow, keywords

    def createLevelMap(self, mergedCells, rSheet):
        """
        将一个 sheet 对象中的合并单元格,转换为层级 map 的过程,为后面树的建立做准备
        :param mergedCells: 合并单元格的range
        :param rSheet: 表对象
        :return: 层级 map, 格式为 level : cellmodel
        """
        # 用于标记表头的最后一行位置
        levellist = []
        setlist = set()
        levelmap = {}
        # 最终降范式后的值
        valuelist = []
        # 这里是如果是 merge cell 的情况, 暂时定 mergecell 范围 > 6 的时候为标题的情况, 所以也要筛除这种情况 TODO
        # 这里是从上往下遍历的 mergecell,所以 mergecells 中 从左到右 row 依次增加,可以利用这点简化计算
        for index, cell in enumerate(mergedCells):
            # 这里要根据 cell 的row 进行分级
            # 根据 sheet 中合并单元格的属性转换为树的节点
            Cmodel = self.createCellModelAccordingRange(cell, rSheet)
            # 去除空值, x 范围一个以上
            if Cmodel.value is not None and Cmodel.value != '':
                levellist.append(Cmodel)
                # 为了找出所有的层级
                setlist.add(Cmodel.level)
        for value in setlist:
            levelmap[value] = []
        for index, cellmodel in enumerate(levellist):
            levelmap[cellmodel.level].append(cellmodel)
        return levelmap

    # 根据sheet 和 遍历的 levelmap 创建节点树
    def createLevelTree(self, levelmap, rSheet):
        """
        根据 levelmap 建立层级树
        :param levelmap: level : 数据此层级的 cellmodel 列表
        :param rSheet: sheet 对象
        :return: 返回树的头结点
        """
        keys = list(levelmap.keys())
        # 最底层的层级
        lastlevel = max(keys)
        minlevel = min(keys)
        # 创建节点树
        # 伪头结点,方便计算
        dummy = JSCellModel()
        dummy.child = levelmap[minlevel]
        # 按层级遍历,将子节点的 x 轴范围数据父节点 x 轴范围的情况进行连接
        for j, curlevel in enumerate(keys):
            # if curlevel == lastlevel:
            currentcells = levelmap[curlevel]
            for i, curcell in enumerate(currentcells):
                crlow, crhign, cclow, cchigh = curcell.cellrange
                if crhign - 1 >= lastlevel or curlevel == lastlevel:
                    # 判断当前是不是最后一层, 且 range > 0的情况
                    if cchigh - cclow > 1:
                        for k in range(cclow, cchigh):
                            cellrange = lastlevel + 1, lastlevel + 2, k, k + 1
                            # print(rSheet.cell_value(lastlevel + 1, i))
                            cmodel = self.createCellModelAccordingRange(cellrange, rSheet)
                            curcell.child.append(cmodel)
                else:
                    # print(curcell.value)
                    nextlevel = keys[j + 1]
                    nextlevelcells = levelmap[nextlevel]
                    for ncell in nextlevelcells:
                        nrlow, nrhign, nclow, nchigh = ncell.cellrange
                        if nclow >= cclow and nchigh <= cchigh:
                            curcell.child.append(ncell)

        self.sortTreeNode(dummy)
        return dummy


    # 对树的每层节点按列进行进行排序
    def sortTreeNode(self, root):
        if (len(root.child) == 0): return
        childs = root.child
        childs = sorted(childs, key=lambda node: node.cellrange[2])
        root.child = childs
        for c in root.child:
            self.sortTreeNode(c)


    # 将 sheet 中的单元格转换为 cellmodel
    def createCellModelAccordingRange(self, cellrange, rSheet):
        """
        :param cellrange:
        :param rSheet: sheet对象
        :return: cellmodel 包括 cell 的值,范围,层级信息
        """
        rlow, rhign, clow, chigh = cellrange
        cellvalue = rSheet.cell(rlow, clow).value
        # self, cellrange=None, value='', level = 1, child=[]
        level = rlow
        # 注意这里要新建,不然会引用同一个 list
        child = []
        Cmodel = JSCellModel(cellrange, cellvalue, level, child)
        return Cmodel


    def travel(self, root, valuelist=[],keywords=[], combinestr='', lv=1):
        """
        树遍历, list 用于接收返回的降范式表头
        :param root:
        :param valuelist:
        :param keywords:
        :param combinestr:
        :param lv:
        :return:
        """
        if root.value == '':
            combinestr = root.value
        else:
            combinestr = "{}/{}".format(combinestr, root.value)
        # 用于判断是否是最后一层
        flag = True
        # print(root.value)
        for nextnode in root.child:
            flag = flag & (len(nextnode.child) == 0)
        # 如果是最后一层
        if flag:
            if len(root.child) == 0:
                combinestr = combinestr.lstrip('/')
                valuelist.append(combinestr)
                keywords.append(root.value)
                return
            for nextlevel in root.child:
                temp = "{}/{}".format(combinestr, nextlevel.value)
                # 由 dummy带来的空字符会在头部多一个/, 剔除
                temp = temp.lstrip('/')
                valuelist.append(temp)
                keywords.append(nextlevel.value)
            return
        else:
            for node in root.child:
                self.travel(node, valuelist, keywords, combinestr, lv)

    # 读取数据表格进行比对后做合并操作 -- 根据表头进行比对来抽取数据
    def ReadDataThenCompareAndExtract(self, configuremodel):
        print("开始抽取合并数据 %s" % datetime.now())
        # 存放表头的路径
        datapath = configuremodel.datapath
        # 读取所有数据文件
        datafilelist = []
        datafilelist = JSExcelHandler.getPathFromRootFolder(datapath, datafilelist)
        if len(datafilelist) == 0:
            messagebox.showinfo("message", "未找到原始数据")
        else:
            # 存放合并结果的路径
            resultfilepath = configuremodel.storepath
            # 类型是 map {路径 : df}, 这里是按照设置的起始行读取标准模板表头,默认是 0
            headsmaplist = self.readStandardHeadFromFolder(configuremodel)
            if headsmaplist == {}:
                messagebox.showinfo("message", "未找到模板表头")
            else:
                for dfpath in headsmaplist.keys():
                    sum = []
                    # 拼接结果文件路径, 用 os.path win 和 linux 会有自己判断
                    filename = os.path.split(dfpath)[-1]
                    startrow = 0
                    rangemap = configuremodel.validrange

                    # 配置模板的起始降范式位置
                    if rangemap != {}:
                        if filename in configuremodel.validrange:
                            startrow = int(configuremodel.validrange[filename])
                    newfile = os.path.join(resultfilepath, filename)
                    '''
                    这里是为所有的标准模板预先建立数据抽取的结果文件
                    '''
                    try:
                        # 降低表头的范式
                        valuelist, lastrows, keywords = self.lowerDimensionOfTitle(dfpath, startrow)
                        # 创建文件及表头文件到result文件夹下为抽取该模板做准备
                        newfile, newsheetname = self.newfilesave(dfpath, newfile, valuelist, startrow)
                    except Exception as e:
                        JSExcelHandler.errorlog("降范式并且新建表头文件未抽取数据做准备-{}".format(dfpath))
                        pass
                    # 遍历目标数据文件下所有 sheet 是否与标准模板匹配,如果匹配则进行数据抽取合并操作
                    # 对文件进行分组合并
                    # def add(work, lock):
                    #     for path in work:
                    #         print(path)
                    #         try:
                    #             readOpenXlsx, sheetnames, tempPath = JSExcelHandler().OpenXls(path)
                    #             for sheetname in sheetnames:
                    #                 # print("校验模板表：" + str(dfpath) + "\n与数据表：" + path + '下工作表sheetname：' + sheetname + "对比")
                    #                 # 这里不能单纯的将标题范围下的数据一并写入,要重新建立按 keywords 排列顺序的 dataframe
                    #                 # 获取模板表头的行数,用于数据表中获取表头范围
                    #                 totalrows = headsmaplist[dfpath].shape[0]
                    #                 # 根据标准表头获取数据里的表头进行比对
                    #                 titledf = pd.read_excel(path, sheet_name=sheetname, nrows=totalrows, skiprows=startrow,
                    #                                         header=None)
                    #                 # 这里要进行最下层 node 的比对
                    #
                    #                 if titledf.equals(headsmaplist[dfpath]):
                    #                     # 抽取模块headsmaplist, 并对匹配上的表进行记录
                    #                     JSExcelHandler.excuteCheck("{} ---- {}".format(dfpath, path))
                    #                     # 抽取数据 dafaframe
                    #                     df = pd.read_excel(path, sheetname, skiprows=lastrows, header=None, dtype='str')
                    #                     filename = os.path.split(path)[-1]
                    #                     df[r'来源文件'] = filename
                    #                     lock.acquire()
                    #                     # self.apendDataFrame(df, newfile, newsheetname, False)
                    #                     sum.append(df)
                    #                     lock.release()
                    #
                    #         except Exception as e:
                    #             print(str(e))
                    #             JSExcelHandler.errorlog(r"遍历目标数据文件下所有 sheet 与标准模板匹配后追加数据-{}".format(path))
                    #
                    # threads = []
                    # lock = threading.Lock()
                    # for i in range(0, len(datafilelist), 5):
                    #     work = datafilelist[i:i + 5]
                    #     # results.append(None) # enlarge the results list, so we have room for this thread's result
                    #     t = threading.Thread(target=add, args=(work, lock ))
                    #     t.start()
                    #     threads.append(t)
                    #
                    # for t in threads:
                    #     t.join()

                    for path in datafilelist:
                        print(path)
                        try:
                            readOpenXlsx, sheetnames, tempPath = JSExcelHandler().OpenXls(path)
                            for sheetname in sheetnames:
                                # print("校验模板表：" + str(dfpath) + "\n与数据表：" + path + '下工作表sheetname：' + sheetname + "对比")
                                # 这里不能单纯的将标题范围下的数据一并写入,要重新建立按 keywords 排列顺序的 dataframe
                                # 获取模板表头的行数,用于数据表中获取表头范围
                                totalrows = headsmaplist[dfpath].shape[0]
                                # 根据标准表头获取数据里的表头进行比对
                                titledf = pd.read_excel(path, sheet_name=sheetname, nrows=totalrows, skiprows=startrow, header=None)
                                # titledf.fillna(method='pad', inplace=True)
                                # 这里要进行最下层 node 的比对
                                if titledf.equals(headsmaplist[dfpath]):
                                    # 抽取模块headsmaplist, 并对匹配上的表进行记录
                                    JSExcelHandler.excuteCheck("{} ---- {}".format(dfpath, path))
                                    # 抽取数据 dafafram
                                    df = pd.read_excel(path, sheetname, skiprows=lastrows, header=None)
                                    filename = os.path.split(path)[-1]
                                    df[r'来源文件'] = '''%s|%s''' % (filename, sheetname)
                                    sum.append(df)
                        except Exception as e:
                            JSExcelHandler.errorlog("降范式并且新建表头文件未抽取数据做准备-{}".format(dfpath))
                            pass
                    res = pd.DataFrame([])
                    for tmpdf in sum:
                       res = res.append(tmpdf)
                    self.apendDataFrame(res, newfile, newsheetname, False)
                print("合并数据完成 %s" % datetime.now())
                messagebox.showinfo("message", "合并完成\n输出路径为:%s" % (configuremodel.storepath))

    # 获取标准的表头的文件,建立标准表头格式的 maplist
    def readStandardHeadFromFolder(self, configuremodel):
        """
        从标准表头存放路径遍历出所有的模板,生成 文件名:模板 dataframe 的 map
        :param configuremodel: view 层的 model
        :return: 标准表头文件名 : 标准表头 dataframe
        """
        headspath = configuremodel.headspath
        headlist = []
        headlist = JSExcelHandler.getPathFromRootFolder(headspath, headlist)
        dfmaplist = {}
        dflist = list(dfmaplist.keys())
        # 从标准头路径中读取标准表头的格式,为比对做准备
        for excelpath in headlist:
            filename = os.path.split(excelpath)[-1]
            startrow = 0
            if filename in configuremodel.validrange:
                startrow = int(configuremodel.validrange[filename])
            # try:
            # path, sheet_name=sheetname, nrows=totalrows, skiprows=startrow, header=None
            # titledf = pd.read_excel(path, sheet_name=sheetname, nrows=totalrows, skiprows=startrow, header=None)
            newdf = pd.read_excel(str(excelpath), skiprows=startrow, header=None)
            # newdf.fillna(method='pad', inplace=True)
            # except Exception as e:
            #     JSExcelHandler.errorlog("获取标准表头-{}".format(excelpath))
            #     pass
            if len(dflist) > 0:
                for olddf in dflist:
                    if newdf.equals(olddf) == True:
                        break
                    else:
                        dfmaplist[excelpath] = newdf
                        dflist = list(dfmaplist.keys())
            else:
                dfmaplist[excelpath] = newdf
                dflist = list(dfmaplist.values())
        return dfmaplist

    # 追加的 dataframe, 目标文件路径, 目标文件 sheet, 是否是 header 类型
    def apendDataFrame(self, apenddf, resultfile, resultsheet, header):
        """
        向目标文件追加数据
        :param apenddf: 追加的 dataframe
        :param resultfile: 目标文件路径
        :param resultsheet: 目标文件的 sheet 名
        :param header: 是否需要保留 header TODO
        """
        # 写入的文件dataframe，engine：以操作工具包执行、mode必须为读状态否者会新增sheet
        writer = pd.ExcelWriter(resultfile, engine='openpyxl')
        book = load_workbook(resultfile)
        writer.book = book
        # 创建 sheets
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        # 获取追加行数
        try:
            dfNum = pd.DataFrame(pd.read_excel(resultfile, sheet_name=resultsheet))
        except Exception as e:
            print(str(e))
            JSExcelHandler.errorlog("追加合并数据-{}".format(resultfile))
            # TODO 这里要进行列的判断
        newRowsNum = dfNum.shape[0] + 1
        # 写入文件
        apenddf.to_excel(excel_writer=writer, sheet_name=resultsheet, index=False, header=header, startrow=newRowsNum)
        writer.save()

    # 新建合并文件
    # oldfile 模板表头文件路径, newfile 合并后数据文件路径
    # 增加是否保留检索的 header
    def newfilesave(self, oldfile, newfile, valuelist, startrow=0, headers=False):
        """
        :param oldfile:
        :param newfile:
        :param valuelist:
        :param startrow:
        :param headers:
        :return:
        """
        readOpenXlsx, sheetnames, tempPath = JSExcelHandler().OpenXls(oldfile)
        # 如果有 startrow 则表明有标题,暂时没有生成标题
        if startrow > 0:
            headers = True
        if headers:
            title = pd.read_excel(oldfile, header=None, nrows=startrow)
            xlwter, sheetnames = self.restoreHead(oldfile, sheetnames[0], newfile, title, startrow)
            df = pd.DataFrame(columns=valuelist)
            df.to_excel(newfile, sheet_name=sheetnames[0], index=False)
            # self.apendDataFrame(df, newfile, sheetnames[0], headers)
        else:
            df = pd.DataFrame(columns=valuelist)
            df.to_excel(newfile, sheet_name=sheetnames[0], index=False)
        return newfile, sheetnames[0]

    # 按照原表头格式还原表格表头, mergecell 的范围重写入表头并按原表合并单元格
    def restoreHead(self, oldfilepath, sheetname, newfile, title, startrow=0):
        """
        因为根据模板表头进行匹配,所以将表头与数据部分分开来,先生成表头部分,再将数据部分进行拼接一次性写入,减少 IO操作
        :param oldfilepath:
        :param sheetname:
        :param newfile:
        :param title:
        :param startrow:
        :return:
        """
        readOpenXlsx, sheetnames, tempPath = JSExcelHandler().OpenXls(oldfilepath)
        writer = pd.ExcelWriter(newfile, engine='xlsxwriter')
        workbook = writer.book
        merge_format = workbook.add_format({'align': 'center'})
        title.to_excel(writer, sheet_name=sheetnames[0], index=False, header=False)
        rsheet = readOpenXlsx.sheet_by_name(sheetname)
        resultCells = rsheet.merged_cells
        mergedCells = []
        # 根据原表中的 cellrange 确定合并单元格的范围, 筛选出 header 部分有合并单元格的情况
        if startrow > 0:
            for index, cell in enumerate(resultCells):
                crlow, crhign, cclow, cchigh = cell
                if crlow < startrow:
                    mergedCells.append(cell)
        else:
            mergedCells = resultCells

        for cellrange in mergedCells:
            crlow, crhign, cclow, cchigh = cellrange
            if crlow < startrow:
                worksheet = writer.sheets[sheetnames[0]]
                value = rsheet.cell_value(crlow, cclow)
                """
                        Merge a range of cells.
                        Args:
                            first_row:    The first row of the cell range. (zero indexed).
                            first_col:    The first column of the cell range.
                            last_row:     The last row of the cell range. (zero indexed).  
                            last_col:     The last column of the cell range.
                            data:         Cell data.
                            cell_format:  Cell Format object.
                """
                # print("{}-{}-{}-{}-{}".format(crlow, cclow, crhign - 1, cchigh - 1, value))
                worksheet.merge_range(crlow, cclow, crhign - 1, cchigh - 1, value, merge_format)
        writer.save()
        return writer, sheetnames

if __name__ == '__main__':

    # GUI入口
    controller = JSETLController()
    tkinter.mainloop()

    # 快速生成模板文件
    # controller = JSETLController()
    # controller.CheckSheetHead(r'C:\Users\sjj-001\Desktop\AUTOETLTEST\heads', r'C:\Users\sjj-001\Desktop\AUTOETLTEST\standard', 1)

    # 合并去重
    # sourcepath = r'C:\Users\sjj-001\Desktop\AUTOETLTEST\result\去重合表目录'
    # objectpath = r'C:\Users\sjj-001\Desktop\AUTOETLTEST\result'
    # JSExcelHandler.dropDuplicatedData(sourcepath, objectpath, 1)

    # 辅助核查
    # exlist = JSExcelHandler.getPathFromRootFolder(r'C:\Users\sjj-001\Desktop\AUTOETLTEST\data')
    # for file in exlist:
    #     print(os.path.split(file)[-1])

    # 文本格式配置文件的测试
    # path = os.path.abspath('..') + "/configureFile.txt"  # 表示当前所处的文件夹上一级文件夹的绝对路径
    # View = JSConfigureView(path)
    # # view 层读取配置文件
    # View.readConfigureFile()
    # # Controller 调度执行读取模板表头文件 | 根据数据文件自动识别表头
    # controller = JSETLController()
    # # 根据标准表头进行比对和抽取合并数据
    # controller.ReadDataThenCompareAndExtract(View.configuremodel)
    # # print("ETL Finished")
    #
    # # 识别数据表中的表头,并抽取表头作为标准表头模板参考
    # controller.autoExtractSheetHead(View.configuremodel)

    # ##处理规则表头和多行表头情况测试
    #
    # # path = 'D:\\\mergeTable\\path\\heads'
    # path = '/Users/sun/Desktop/AUTOEtlTest/heads/'
    # list = JSExcelHandler.getPathFromRootFolder(path)
    # controller = JSETLController()
    # for path in list:
    #     readOpenXls, sheetnames, workpath = JSExcelHandler.OpenXls(path)
    #     for sheetname in sheetnames:
    #         rSheet = readOpenXls.sheet_by_name(sheetname)
    #         controller.lowerDimensionOfTitle(rSheet, 2)

    # controller.newfilesave('D:\\\mergeTable\\path\\result\\result.xlsx',['id', 'nmae'])




